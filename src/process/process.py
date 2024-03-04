# resize image
# process_single_image
# process_form_folder

import cv2
import time
import numpy as np
import matplotlib.pyplot as plt
import os
import pandas as pd
from openpyxl import load_workbook

def resise_image(image, scale_percent=40):
    """Resize an image to a given scale."""
    #width = int(image.shape[1] * scale_percent / 100)
    #height = int(image.shape[0] * scale_percent / 100)

    width = 1970
    height = 1436
    dim = (width, height)
    resized_image = cv2.resize(image, dim)
    return resized_image

def process_single_image(image_path, superpoint, display=False, resize=40):
    """Process an individual image to extract keypoints and descriptors."""

    print(f"==> Processing image: {image_path}...")
    start_time = time.time()
    image = cv2.imread(image_path, 0)
    image = np.float32(image) / 255.0
    
    resised_image = resise_image(image, resize)
    keypoints, descriptors, _ = superpoint.run(resised_image)
        
    if display:
        plt.figure(figsize=(10,10))
        plt.imshow(resised_image, cmap='gray')
        plt.scatter(keypoints[0,:], keypoints[1,:], color='r', s=3)
        plt.show()
        
    print(f"==> Finished processing image. Time taken: {time.time() - start_time:.2f} seconds.")
    return keypoints, descriptors

def process_form_folder(form_folder_path, superpoint):
    """Process all images within a folder."""

    print(f"==> Processing form folder: {form_folder_path}...")
    start_time = time.time()
    keypoints_list, descriptors_list = [], []
    desc_ref = None
    kp_ref = None
    
    folder_name = os.path.basename(form_folder_path)
    
    for image_name in os.listdir(form_folder_path):
        image_path = os.path.join(form_folder_path, image_name)
        
        # Check if the image has the same name as the folder (without file extension)
        name_without_extension, _ = os.path.splitext(image_name)
        
        if image_path.lower().endswith(('.png', '.jpg', '.tif')):
            keypoints, descriptors = process_single_image(image_path, superpoint)
            
            print(folder_name)
            print(name_without_extension)
            if name_without_extension == folder_name:
                print('save ref')
                desc_ref = descriptors
                kp_ref = keypoints
            else:
                keypoints_list.append(keypoints)
                descriptors_list.append(descriptors)
    
    print(f"==> Finished processing form folder. Time taken: {time.time() - start_time:.2f} seconds.")
    
    return keypoints_list, descriptors_list, desc_ref, kp_ref

def choose_good_excel(match_form):
    if '2042_K_' in match_form:
        return '2042K'
    elif '2042_Kauto' in match_form:
        return '2042KAUTO'
    else:
        return '2042'
    
def display_images(img1, img2, title1='Image 1', title2='Image 2'):
    """
    Display two images side by side using matplotlib

    :param img1: First image to display
    :param img2: Second image to display
    :param title1: Title of the first image
    :param title2: Title of the second image
    """
    
    fig, axs = plt.subplots(1, 2, figsize=(20, 10))
    
    axs[0].imshow(img1, cmap='gray')
    axs[0].axis('off')
    axs[0].set_title(title1)
    
    axs[1].imshow(img2, cmap='gray')
    axs[1].axis('off')
    axs[1].set_title(title2)
    
    plt.tight_layout()
    plt.show()

def append_df_to_excel(df, excel_path, sheet_name):
    workbook = load_workbook(excel_path)
    sheet = workbook[sheet_name]
    
    max_row_by_col = {}
    for column in df:
        max_row_by_col[column] = sheet.max_row
    
    # Obtenir les valeurs à ajouter depuis le DataFrame
    for column in df:
        # Obtenir la première rangée vide pour la colonne
        row = max_row_by_col[column] + 1
        # Obtenir la valeur à ajouter - on suppose qu'il y a une seule valeur par colonne dans le DataFrame
        value = df.at[0, column]  # Obtient la valeur de la première rangée pour la colonne actuelle
        if pd.notna(value):  # Vérifier si la valeur n'est pas NaN
            # Mettre à jour la cellule dans le fichier Excel
            sheet[f"{column}{row}"] = value
    
    # Sauvegarder le classeur
    workbook.save(excel_path)

def create_paths(df):
    """
    Creates a dictionary where each key is a concatenated path from 'Lot' and 'Image' columns
    and each value is the corresponding 'SPI' value, filtering out rows where 'SPI' is '0000000000000'.
    
    :param df: DataFrame
        Input DataFrame to be processed.
    :return: dict
        A dictionary containing paths as keys and 'SPI' values as values.
    """
    paths_dict = {}
    try:
        # Iterating over rows in the DataFrame
        for index, row in df.iterrows():
                path = os.path.join(str(row['Lot']), str(row['Image']))
                paths_dict[index] = path

                
    except KeyError as e:
        print(f"Error: Column not found in the DataFrame: {e}")
    except Exception as e:
        print(f"An error occurred while processing the DataFrame: {e}")
        
    return paths_dict

def load_file_into_dataframe(file_path):
    """
    Load an Excel or ODS file into a DataFrame.
    
    :param file_path: str
        Path to the file to be loaded.
    :return: DataFrame or None
        Returns a DataFrame if successful, otherwise None.
    """
    try:
        # Loading Excel file
        if file_path.endswith('.xlsx'):
            df = pd.read_excel(file_path, engine='openpyxl')
        # Loading ODS file
        elif file_path.endswith('.ods'):
            df = pd.read_excel(file_path, engine='odf')
        else:
            print("Error: Unsupported file format. Please provide an .ods or .xlsx file.")
            return None
        
        return df
    
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
    except PermissionError:
        print(f"Error: Permission denied to access the file '{file_path}'.")
    except Exception as e:
        print(f"An error occurred while loading the file: {e}")

def get_paths_dict(path=None, verbose=False):

    if path is None:
        path = '../../Data/POC/table_correspondance_90pourcents.ods'

    df = load_file_into_dataframe(path)

    paths_dict = create_paths(df)

    if verbose:
        print(len(paths_dict))
        for spi, path in paths_dict.items():
            print(f"SPI: {spi}, Path: {path}")
    
    return paths_dict